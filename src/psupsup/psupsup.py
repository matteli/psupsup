import json
import zipfile
import os
import datetime
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill
from fpdf import FPDF
from transformers import pipeline


def initialiser_classificateur(model_path="./models/appert", device=0):
    """
    Initialise le modèle de classification d'appréciations.

    ### Args:
    - model_path (str, optional): Le chemin du modèle à utiliser (par défaut "./models/appert").
    - device (int, optional): Le numéro du dispositif à utiliser (par défaut 0 : 1er GPU, -1 : CPU).

    ### Returns:
    - pipeline: Un pipeline de classification de texte initialisé avec le modèle spécifié.
    """
    print("Initialisation du modèle de classification d'appréciations...")
    return pipeline(
        "text-classification", model=model_path, device=device, local_files_only=True
    )


def classer_app(classifier, appreciation):
    """
    Classifie une appréciation (0 : mauvaise appréciation, 1 : bonne appréciation) en utilisant le modèle de classification.

    ### Args:
    - classifier (pipeline): Un pipeline de classification de texte.
    - appreciation (str | list): La ou les appréciations à classer.

    ### Returns:
    - list: Le résultat de la classification de l'appréciation, avec les étiquettes et les scores correspondants dans un dictionnaire.
    """
    if appreciation is not None:
        cl = classifier(appreciation)
        return cl[0]["label"]
    return None


def charger_json(chemin_fichier):
    """
    Charge un fichier JSON à partir d'un fichier ZIP et retourne les données sous forme de dictionnaire.

    ### Args:
    - chemin_fichier (str): Le chemin du fichier ZIP contenant le JSON à charger.

    ### Returns:
    - dict: Un dictionnaire contenant les données du JSON, ou None en cas d'erreur.
    """
    try:
        with zipfile.ZipFile(chemin_fichier, "r") as zip_ref:
            with zip_ref.open(zip_ref.namelist()[0]) as fichier_json:
                # with open(fichier_json.read(), "r", encoding="utf-8") as fichier:
                donnees = json.load(fichier_json)
                return donnees
    except FileNotFoundError:
        print(f"Erreur : Le fichier '{chemin_fichier}' est introuvable.")
    except zipfile.BadZipFile:
        print(
            f"Erreur : Le fichier '{chemin_fichier}' n'est pas un fichier ZIP valide."
        )
    except json.JSONDecodeError:
        print(
            f"Erreur : Le fichier '{chemin_fichier}' dézippé n'est pas un JSON valide."
        )
    return None


def selectionner_fichier(dossier, extension):
    """
    Affiche un menu avec une liste d'options et retourne le choix de l'utilisateur.

    ### Args:
    - dossier (str): Le chemin du dossier à lister.
    - extension (str): L'extension des fichiers à lister.

    ### Returns:
    - str: L'option choisie par l'utilisateur ou le fichier présent s'il n'y en a qu'un, ou None si une erreur survient.
    """
    fichiers = []
    try:
        for fichier in os.listdir(dossier):
            if fichier.endswith(extension):
                fichiers.append(fichier)
    except FileNotFoundError:
        print(f"Erreur : Le dossier '{dossier}' est introuvable.")

    if len(fichiers) == 0:
        print(f"Aucun fichier .{extension} trouvé dans le dossier.")
        exit()
    elif len(fichiers) == 1:
        print(f"Un seul fichier trouvé : {fichiers[0]}")
        return fichiers[0]
    print("\n--- Fichiers disponibles ---")
    for i, option in enumerate(fichiers, 1):
        print(f"{i}. {option}")

    while True:
        try:
            choix = int(input("\nEntrez votre choix (numéro) : "))
            if 1 <= choix <= len(fichiers):
                return fichiers[choix - 1]
            else:
                print(f"Erreur : Veuillez entrer un numéro entre 1 et {len(fichiers)}.")
        except ValueError:
            print("Erreur : Veuillez entrer un nombre valide.")


def charger_donnees(fichier=None):
    """
    Charge les données d'un fichier JSON ZIP et retourne la liste des candidats.

    ### Args:
    - fichier (str, optional): Le chemin du fichier ZIP contenant le JSON à charger. Si None, affiche un menu pour choisir un fichier.

    ### Returns:
    - list: Une liste de candidats, ou None en cas d'erreur.
    """
    if fichier is None:
        fichier = selectionner_fichier(".", ".json.zip")
    dictionnaire = charger_json(fichier)
    if dictionnaire is not None:
        candidats = dictionnaire["exportDeDonnees"]["exportCandidats"][0]["candidats"]
        return candidats
    return None


def convertir_chaine_en_float(chaine):
    """
    Convertit une chaîne contenant un nombre décimal avec des virgules en un float.

    ### Args:
    - chaine (str): La chaîne à convertir.

    ### Returns:
    - float: Le nombre converti en float, ou None si la conversion échoue.
    """
    try:
        return float(chaine.replace(",", "."))
    except (ValueError, AttributeError):
        return None


def calculer_note_modifiee(
    note, moyenne_classe, moyenne_basse, moyenne_haute, brute=0.5
):
    """
    Calcule une note modifiée en pondérant avec l'argument 'brute' la note brute et la note normalisée.

    ### Args:
    - note (any): La note à modifier.
    - moyenne_classe (any): La moyenne de la classe.
    - moyenne_basse (any): La moyenne la plus basse.
    - moyenne_haute (any): La moyenne la plus haute.
    - brute (float, optional): Le poids de la note brute (entre 0 et 1, par défaut 0.5).

    ### Returns:
    - float: La note modifiée, ou -1 si une erreur survient.
    """
    note = convertir_chaine_en_float(note)
    if note is None:
        return -1
    moyenne_classe = convertir_chaine_en_float(moyenne_classe)
    if moyenne_classe is None:
        return note / 2
    moyenne_basse = convertir_chaine_en_float(moyenne_basse)
    if moyenne_basse is None:
        moyenne_basse = moyenne_classe / 2
    moyenne_haute = convertir_chaine_en_float(moyenne_haute)
    if moyenne_haute is None:
        moyenne_haute = 20

    return (
        calculer_note_normalisee(note, moyenne_classe, moyenne_basse, moyenne_haute)
        * (1 - brute)
        + note * brute
    )


def calculer_note_normalisee(note, moyenne_classe, moyenne_basse, moyenne_haute):
    """
    Calcule la regression linéaire entre 0 et 20 pour une note avec la moyenne de la classe en valeur centrale.

    ### Args:
    - note (float): La note à normaliser.
    - moyenne_classe (float): La moyenne de la classe.
    - moyenne_basse (float): La moyenne la plus basse.
    - moyenne_haute (float): La moyenne la plus haute.

    ### Returns:
    - float: La note normalisée, ou -1 si une erreur survient.
    """
    if note < moyenne_classe:
        if moyenne_classe == moyenne_basse:
            return 10
        return (note - moyenne_basse) / (moyenne_classe - moyenne_basse) * 10
    else:
        if moyenne_classe == moyenne_haute:
            return 10
        return 10 + (note - moyenne_classe) / (moyenne_haute - moyenne_classe) * 10


def convertir_note_bac(note):
    """
    Convertit une note de baccalauréat en float et retourne 0 si la conversion échoue.

    ### Args:
    - note (str): La note à convertir.

    ### Returns:
    - float: La note convertie en float, ou 0 si la conversion échoue.
    """
    note = convertir_chaine_en_float(note)
    if note is None:
        return 0
    return note


def candidature_confirmee(candidat):
    """
    Vérifie si la candidature d'un candidat est confirmée.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.

    ### Returns:
    - bool: True si la candidature est confirmée, False sinon.
    """
    return int(candidat["DonneesVoeux"]["CandidatureConfirmeeCode"]) == 1


def numero_dossier_candidat(candidat):
    """
    Retourne le numéro de dossier d'un candidat.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.

    ### Returns:
    - str: Le numéro de dossier du candidat.
    """
    return candidat["DonneesCandidats"]["NumeroDossierCandidat"]


def sexe_candidat(candidat):
    """
    Retourne le sexe d'un candidat.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.

    ### Returns:
    - str: Le sexe du candidat (Masculin ou Féminin).
    """
    return candidat["DonneesCandidats"].get("Sexe", "N/A")


def type_bac(candidat):
    """
    Retourne le type de baccalauréat d'un candidat, en distinguant les bacs professionnels aéronautiques.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.

    ### Returns:
    - dict: La série (serie) (Générale, STI2D, P, STL, STMG...) et la spécialité (specialite) (Energies et environnement, Innovation technologique et eco conception, Architecture et construction, Système informatique et numérique, Aéronautique opt. mécanicien syst. avionique...) du baccalauréat du candidat.
    """

    return {
        "serie": candidat["Baccalaureat"]["SerieDiplomeCode"],
        "specialite": candidat["Baccalaureat"].get("SpecialiteLibelle", "N/A"),
    }


def identite_candidat(candidat):
    """
    Retourne un dictionnaire contenant le nom, le prénom et l'email d'un candidat.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.

    ### Returns:
    - dict: Un dictionnaire contenant le nom (nom), le prénom (prenom) et l'email (email) du candidat.
    """
    return {
        "nom": candidat.get("DonneesCandidats", {}).get("NomCandidat", "N/A"),
        "prenom": candidat.get("DonneesCandidats", {}).get("PrenomCandidat", "N/A"),
        "email": candidat.get("DonneesCandidats", {}).get(
            "CoordonneesAdressemail", "N/A"
        ),
    }


def groupe_candidat(candidat):
    """
    Retourne le groupe d'un candidat.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.

    ### Returns:
    - str: Le groupe du candidat (Autres candidats ou Bacheliers professionnels toutes séries), ou "N/A" si l'information n'est pas disponible.
    """
    return candidat.get("DonneesVoeux", "N/A").get("GroupeLibelle", "N/A")


def filtrer_scolarite(candidat, redoublement_neutralise, terminale_seulement):
    """
    Retourne une liste des années scolaires lycée du candidat en filtrant les redoublements de Première et Terminale si 'redoublement_neutralise' est True.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.
    - redoublement_neutralise (bool): Si True, les redoublements de Première et Terminale sont neutralisés.
    - terminale_seulement (bool): Si True, les années de Terminale sont retournées.

    ### Returns:
    - list: Une liste des années scolaires lycée du candidat, filtrée selon les redoublements si nécessaire.
    """
    sco_candidat = []
    T1 = False
    P1 = False
    for sco in candidat["Scolarite"]:
        if sco.get("NiveauEtudeLibelle") == "Terminale":
            if not T1 or not redoublement_neutralise:
                sco_candidat.append(sco.get("AnneeScolaireCode"))
                T1 = True
        elif sco.get("NiveauEtudeLibelle") == "Première":
            if not P1 or not redoublement_neutralise or not terminale_seulement:
                sco_candidat.append(sco.get("AnneeScolaireCode"))
                P1 = True
    return sco_candidat


def iterer_matieres_dans_bac(candidat, notes_bac):
    """
    Itère sur les matières du baccalauréat d'un candidat et retourne un dictionnaire contenant la matière, le code et la note.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.
    - notes_bac (dict): Un dictionnaire associant les matières du baccalauréat à leurs codes.

    ### Returns:
    - dict: Un dictionnaire contenant la matière, le code et la note pour chaque matière trouvée dans les notes de baccalauréat du candidat.
    """
    for NoteBac in candidat["NotesBaccalaureat"]:
        for matiere, codes in notes_bac.items():
            for c in codes:  # matiere:
                if c == int(NoteBac["EpreuveCode"]):
                    yield {
                        "matiere": matiere,
                        "code": c,
                        "note": NoteBac.get("NoteEpreuve", 0),
                    }


def iterer_matieres_dans_bulletins(
    candidat, matieres, redoublement_neutralise, terminale_seulement
):
    """
    Itère sur les matières des bulletins scolaires filtrés (redoublement ou terminale) d'un candidat et retourne un dictionnaire contenant la matière, le code, les moyennes et l'appréciation.

    ### Args:
    - candidat (dict): Un dictionnaire contenant les données d'un candidat.
    - matieres (dict): Un dictionnaire associant les matières à leurs codes.
    - redoublement_neutralise (bool): Si True, les redoublements de Première et Terminale sont neutralisés.
    - terminale_seulement (bool): Si True, les années de Terminale sont retournées.

    ### Returns:
    - dict: Un dictionnaire contenant la matière, le code, les moyennes du candidats, de classe, haute et basse et l'appréciation pour chaque matière trouvée dans les bulletins scolaires du candidat.
    """
    sco_candidat = filtrer_scolarite(
        candidat, redoublement_neutralise, terminale_seulement
    )
    for annee_code in sco_candidat:
        for BulletinScolaire in candidat["BulletinsScolaires"]:
            if BulletinScolaire["AnneeCode"] == annee_code:
                for BulletinsScolairesSeries in BulletinScolaire[
                    "BulletinsScolairesAnnee"
                ]["BulletinsScolairesSeries"]:
                    for BulletinMatiere in BulletinsScolairesSeries[
                        "BulletinsScolairesParPeriode"
                    ]:
                        for matiere, codes in matieres.items():
                            for c in codes:  # matiere:
                                if c == 0 or c == int(
                                    BulletinMatiere["MatiereBulletinCode"]
                                ):
                                    yield {
                                        "matiere": matiere,
                                        "code": c,
                                        "moyenne_candidat": BulletinMatiere.get(
                                            "MoyenneduCandidat", "N/A"
                                        ),
                                        "moyenne_classe": BulletinMatiere.get(
                                            "MoyenneclasseCandidat", "N/A"
                                        ),
                                        "moyenne_basse": BulletinMatiere.get(
                                            "MoyenneBasseClasseduCandidat", "N/A"
                                        ),
                                        "moyenne_haute": BulletinMatiere.get(
                                            "MoyenneHauteClasseduCandidat", "N/A"
                                        ),
                                        "appreciation": BulletinMatiere.get(
                                            "AppreciationProfesseur"
                                        ),
                                    }


def classer_candidats_par_critère(resultats_candidats, critere):
    """
    Retourne un dictionnaire de résultats de candidats triés par un critère donné en ordre décroissant.

    ### Args:
    - resultats_candidats (dict): Un dictionnaire contenant les résultats des candidats.
    - critere (str): Le critère de tri (par exemple, "Notes").

    ### Returns:
    - dict: Un dictionnaire de résultats de candidats triés par le critère donné en ordre décroissant.
    """
    return dict(
        sorted(
            resultats_candidats.items(),
            key=lambda x: x[1][critere],
            reverse=True,
        )
    )


def generer_excel(resultats_candidats, headers, nom_fichier="resultats_candidats.xlsx"):
    """
    Génère un fichier Excel à partir d'un dictionnaire de résultats de candidats, en ajoutant les nouveaux candidats à un fichier existant si nécessaire.

    ### Args:
    - resultats_candidats (dict): Un dictionnaire contenant les résultats des candidats.
    - headers (list): Une liste des en-têtes de colonnes à inclure dans le fichier Excel.
    - nom_fichier (str, optional): Le nom du fichier Excel à générer (par défaut "resultats_candidats.xlsx").

    ### Returns:
    - bool: True si le fichier Excel a été généré avec succès, False sinon.
    """
    list_num = []
    if os.path.exists(nom_fichier):
        wb = openpyxl.load_workbook(nom_fichier)
        ws = wb["Résultats"]
        list_num = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]
        for num, candidat_data in resultats_candidats.items():
            if num in list_num:
                for i, h in enumerate(headers):
                    if h not in candidat_data:
                        candidat_data[h] = ws.cell(
                            row=list_num.index(num) + 2, column=i + 2
                        ).value
        os.rename(
            nom_fichier,
            nom_fichier.replace(
                ".xlsx",
                f"_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            ),
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    ws.append(["Num"] + headers)

    # Style headers
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    add = False
    for num, candidat_data in resultats_candidats.items():
        ws.append([num] + [candidat_data.get(d, "") for d in headers])
        if num not in list_num:
            print(f"Ajouté: {num}")
            add = True
    if not add:
        print("Aucun candidat n'a été ajouté.")

    # Auto-fit columns
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    wb.save(nom_fichier)
    return True


def generer_pdf(
    resultats_candidats,
    titre,
    headers,
    col_width,
    nom_fichier,
    criteres=[],
):
    """
    Génère un fichier PDF à partir d'un dictionnaire de résultats de candidats, en filtrant les candidats selon des critères d'inclusion ou d'exclusion.

    ### Args:
    - resultats_candidats (dict): Un dictionnaire contenant les résultats des candidats.
    - titre (str): Le titre à afficher en haut du PDF.
    - headers (list): Une liste des en-têtes de colonnes à inclure dans le PDF.
    - col_width (list): Une liste des largeurs de colonnes correspondantes aux en-têtes.
    - nom_fichier (str): Le nom du fichier PDF à générer.
    - criteres (list, optional): Une liste de dictionnaires contenant les critères d'inclusion ou d'exclusion pour filtrer les candidats (par exemple, [{"critere": "Bac", "inclus": ["P"]}]).

    ### Returns:
    - bool: True si le PDF a été généré avec succès, False sinon.
    """

    class PDF(FPDF):
        def header(self):
            self.set_font("Courier", "B", 10)
            self.cell(0, 10, titre, border=0, align="C")
            self.ln(10)

            for h in headers:
                self.cell(col_width[(headers).index(h)], 10, h, border=1, align="C")
            self.ln()

        def footer(self):
            self.set_y(-15)
            self.set_font("Courier", "", 10)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Courier", "", 10)

    for num, data in resultats_candidats.items():
        if all(
            data.get(c["critere"]) in c["inclus"] for c in criteres if "inclus" in c
        ):
            if all(
                data.get(c["critere"]) not in c["exclus"]
                for c in criteres
                if "exclus" in c
            ):
                for h in headers:
                    if h == "Num":
                        pdf.cell(col_width[0], 10, str(num), border=1)
                    else:
                        text = data.get(h, "")
                        if isinstance(text, float):
                            text = f"{text:.3f}"
                        elif isinstance(text, int):
                            text = str(text)
                        pdf.cell(
                            col_width[headers.index(h)],
                            10,
                            text,
                            border=1,
                            align="C",
                        )
                pdf.ln()

    pdf.output("./" + nom_fichier)
    return True


def generer_csv(resultats_candidats, colonnes, chemin_fichier=None):
    """
    Génère un fichier CSV importable dans parcoursup à partir d'un dictionnaire de résultats de candidats et d'un fichier csv exporté.

    ### Args:
    - resultats_candidats (dict): Un dictionnaire contenant les résultats des candidats.
    - colonnes (list): Une liste des noms de colonnes à compléter dans le fichier CSV.
    - chemin_fichier (str, optional): Le chemin du fichier CSV exporté de parcoursup.

    ### Returns:
    - bool: True si le CSV a été généré avec succès, False sinon.
    """
    if chemin_fichier is None:
        fichier = selectionner_fichier(".", ".csv")
    else:
        fichier = chemin_fichier

    res = []

    with open(fichier, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        existing_headers = next(reader)
        if existing_headers[:7] != [
            "Code Candidat",
            "Nom candidat",
            "Prénom candidat",
            "Groupe",
            "Code examinateur",
            "Nom examinateur",
            "Prénom examinateur",
        ]:
            print("Erreur : Le fichier CSV sélectionné n'a pas le format attendu.")
            return False
        for row in reader:
            num = row[0]
            if num in resultats_candidats:
                for i, h in enumerate(colonnes):
                    if h in resultats_candidats[num]:
                        row[existing_headers.index(h)] = resultats_candidats[num][h]
            res.append(row)
    os.rename(
        fichier,
        fichier.replace(
            ".csv",
            f"_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.csv",
        ),
    )
    with open(fichier, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(existing_headers)
        for row in res:
            writer.writerow(row)

    return True
